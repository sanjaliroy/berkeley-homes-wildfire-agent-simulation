"""
Preview test for run_evaluation_v2.ipynb — no LLM calls.

Executes the notebook's actual statistics, table, and Excel-export cells against
synthetic judge scores, and writes a sample workbook to:

    outputs/eval/evaluation_v2_claude_SAMPLE.xlsx

Run it any time you change the evaluation notebook to check the output shape:

    python -m pytest tests/test_eval_export_preview.py -s
"""

import json
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

pytest.importorskip("openpyxl")
pytest.importorskip("scipy")

PROJECT = Path(__file__).resolve().parents[1]
NOTEBOOK = PROJECT / "notebooks" / "run_evaluation_v2.ipynb"

DIMS = ["behavioral_plausibility", "persona_consistency", "intervention_responsiveness"]
VARIANTS = ["Baseline", "Ablation1_No_Reflection", "Ablation2_No_Memory_No_Reflection", "Budget"]
AGENTS = [("laura", "Laura"), ("linda", "Linda"), ("walter", "Walter"),
          ("margaret", "Margaret"), ("miriam", "Miriam Voss")]
EVENTS = [(1, "ordinance_notification"), (12, "defensible_space_inspection"),
          (24, "neighbor_pressure"), (36, "firewise_community_campaign"),
          (48, "insurance_nonrenewal_notice"), (55, "resource_assistance_email")]
# Synthetic effect sizes (relative to Baseline) so the tables show plausible deltas
SHIFT = {"Baseline": 0.0, "Ablation1_No_Reflection": -0.05,
         "Ablation2_No_Memory_No_Reflection": -0.15, "Budget": -0.07}


def _cells():
    nb = json.loads(NOTEBOOK.read_text())
    return [c["source"] if isinstance(c["source"], str) else "".join(c["source"])
            for c in nb["cells"] if c["cell_type"] == "code"]


def _cell(cells, marker):
    hits = [s for s in cells if marker in s]
    assert len(hits) == 1, f"marker {marker!r} matched {len(hits)} cells"
    return hits[0]


def _synthetic_frames(ns):
    rng = np.random.default_rng(42)
    fullsim_rows, runs = [], {}
    for v in VARIANTS:
        for rep in range(1, 6):
            label = f"claude_{v}_rep{rep}"
            runs[label] = {
                "variant": v, "rep": rep, "decisions": [], "source_file": f"{label}.jsonl",
                "run_config": {"run_label": label,
                               "decision_model": "claude-haiku-4-5-20251001" if v == "Budget"
                               else "claude-sonnet-4-6"},
                "run_summary": {"agent_cost_usd": round(0.5 + rng.normal(0, 0.05), 4),
                                "latency_seconds": round(500 + rng.normal(0, 30), 1)},
            }
            for aid, name in AGENTS:
                row = {"run_label": label, "run_id": f"{label}_20260710_000000",
                       "variant": v, "rep": rep, "agent_id": aid,
                       "agent_display_name": name, "bp_note": "sample", "pc_note": "sample",
                       "ir_note": "sample", "judge_model": "openai/gpt-5.4"}
                for d in DIMS:
                    row[d] = float(np.clip(4.8 + SHIFT[v] + rng.normal(0, 0.15), 1, 5))
                fullsim_rows.append(row)
    fullsim = pd.DataFrame(fullsim_rows)
    fullsim["overall"] = fullsim[DIMS].mean(axis=1)

    jv_rows = []
    for repeat in range(1, 6):
        for aid, name in AGENTS:
            row = {"run_label": "claude_Baseline_rep1", "run_id": "claude_Baseline_rep1_20260710_000000",
                   "variant": "Baseline", "rep": 1,
                   "agent_id": aid, "agent_display_name": name, "repeat": repeat,
                   "bp_note": "sample", "pc_note": "sample", "ir_note": "sample",
                   "judge_model": "openai/gpt-5.4"}
            for d in DIMS:
                row[d] = float(np.clip(4.8 + rng.normal(0, 0.05), 1, 5))
            jv_rows.append(row)
    judge_var = pd.DataFrame(jv_rows)
    judge_var["overall"] = judge_var[DIMS].mean(axis=1)

    pi_rows = []
    for v in VARIANTS:
        for aid, name in AGENTS:
            for day, event_type in EVENTS:
                row = {"run_label": f"claude_{v}_rep1", "run_id": f"claude_{v}_rep1_20260710_000000",
                       "variant": v, "rep": 1,
                       "agent_id": aid, "agent_display_name": name, "day": day,
                       "event_type": event_type, "bp_note": "sample", "pc_note": "sample",
                       "ir_note": "sample", "judge_model": "openai/gpt-5.4"}
                for d in DIMS:
                    row[d] = float(np.clip(4.7 + SHIFT[v] + rng.normal(0, 0.25), 1, 5))
                pi_rows.append(row)
    interventions = pd.DataFrame(pi_rows)
    interventions["overall"] = interventions[DIMS].mean(axis=1)

    ns.update(fullsim=fullsim, judge_var=judge_var, interventions=interventions,
              runs=runs, jv_label="claude_Baseline_rep1", JV_REP=1)


class _FixedDatetime:
    """Stand-in for datetime so the sample export filename is deterministic."""
    @staticmethod
    def now():
        class _T:
            @staticmethod
            def strftime(_fmt):
                return "SAMPLE"
        return _T()


def test_eval_export_preview(capsys):
    cells = _cells()
    ns = {"PROJECT_PATH": str(PROJECT)}

    exec(_cell(cells, "from src.llm.client import"), ns)      # imports (no API calls)
    exec(_cell(cells, "MODEL_FAMILY = 'claude'"), ns)          # configuration

    _synthetic_frames(ns)

    # All analysis/table cells, in notebook order. Markers must each match exactly
    # one cell — several substrings now appear in both an old cell and the new
    # report-tables cell, so anchor them to the defining statement.
    for marker in ["def mean_se", "base_means = {c:", "MODEL COMPARISON", "COST AND LATENCY",
                   "per_repeat = judge_var.groupby", "FULL SIMULATION SCORES BY AGENT",
                   "SCORING BY CRITERION", "def event_label", "REPORT_TABLES = ["]:
        exec(_cell(cells, marker), ns)

    # Export cell with a deterministic filename
    ns["datetime"] = _FixedDatetime
    exec(_cell(cells, "ExcelWriter"), ns)

    sample_path = PROJECT / "outputs" / "eval" / "evaluation_v2_claude_SAMPLE.xlsx"
    assert sample_path.exists(), sample_path

    import openpyxl
    wb = openpyxl.load_workbook(sample_path, read_only=True)
    expected_sheets = {"Report Tables", "FullSim Scores", "Replicate Level", "Judge Variance",
                       "Per-Intervention", "Cost & Latency",
                       "Appx A FullSim by Agent", "Appx B Baseline Criteria"}
    assert expected_sheets <= set(wb.sheetnames), wb.sheetnames
    assert wb.sheetnames[0] == "Report Tables", f"Report Tables must be first: {wb.sheetnames}"
    assert "Sheet" not in wb.sheetnames, "leftover default 'Sheet' not removed"

    out = capsys.readouterr().out
    for heading in ["6.2.1 OVERALL RESULTS", "6.2.3 MODEL COMPARISON",
                    "6.2.4 COST AND LATENCY", "JUDGE VARIANCE",
                    "FULL SIMULATION SCORES BY AGENT", "CI excludes 0",
                    "REPORT-FORMATTED TABLES", "Simulation Variance (run-to-run)",
                    "Variance Comparison"]:
        assert heading in out, f"missing table heading: {heading}"

    print(f"\nSample workbook written to: {sample_path}")
